"""
data_exporter.py
----------------
Export journey results to CSV, Excel (.xlsx), and JSON.
Designed for Jupyter notebook use with download links.
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _timestamp_str() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: str):
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)


# ---------------------------------------------------------------------------
# SINGLE JOURNEY EXPORT
# ---------------------------------------------------------------------------

def export_journey_csv(result: Dict, output_dir: str = "outputs") -> str:
    """Export single journey leg detail to CSV. Returns file path."""
    from modules.carbon_calculator import results_to_dataframe

    df = results_to_dataframe(result)
    os.makedirs(output_dir, exist_ok=True)
    origin = result["origin_resolved"].split(",")[0].replace(" ", "_")
    dest = result["destination_resolved"].split(",")[0].replace(" ", "_")
    fname = f"{output_dir}/journey_{origin}_to_{dest}_{_timestamp_str()}.csv"
    df.to_csv(fname, index=False)
    print(f"✓ CSV saved: {fname}")
    return fname


def export_journey_xlsx(result: Dict, output_dir: str = "outputs") -> str:
    """Export single journey to Excel with formatted sheets."""
    from modules.carbon_calculator import results_to_dataframe

    if not OPENPYXL_AVAILABLE:
        print("openpyxl not installed. Run: pip install openpyxl")
        return None

    df = results_to_dataframe(result)
    os.makedirs(output_dir, exist_ok=True)
    origin = result["origin_resolved"].split(",")[0].replace(" ", "_")
    dest = result["destination_resolved"].split(",")[0].replace(" ", "_")
    fname = f"{output_dir}/journey_{origin}_to_{dest}_{_timestamp_str()}.xlsx"

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        # Sheet 1: Leg details
        df.to_excel(writer, sheet_name="Journey Detail", index=False)

        # Sheet 2: Summary
        summary_data = {
            "Field": [
                "Origin", "Destination", "Route", "Distance (km)",
                "Cabin Class", "Return Trip",
                "One-way Transport CO2 (kg)", "Total Transport CO2 (kg)",
                "Hotel Nights", "Hotel CO2 (kg)", "GRAND TOTAL CO2 (kg CO2e)",
                "Data Sources", "Calculation Date"
            ],
            "Value": [
                result["origin_resolved"],
                result["destination_resolved"],
                result["route_name"],
                result["total_distance_km"],
                result.get("cabin_class", "economy").title(),
                "Yes" if result["return_trip"] else "No",
                result["one_way_transport_co2_kg"],
                result["total_transport_co2_kg"],
                result["nights_at_destination"],
                result["hotel_co2_kg"],
                result["grand_total_co2_kg"],
                "DEFRA 2025 / Google TIM API / IEA 2023 / EEA 2023",
                result["timestamp"][:10]
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # Sheet 3: All route options
        if result.get("all_route_options"):
            opt_rows = []
            for opt in result["all_route_options"]:
                opt_rows.append({
                    "Route Option": opt["name"],
                    "Total Distance (km)": round(opt.get("total_km", 0), 1),
                    "Total CO2 (kg)": round(opt.get("total_co2_kg", 0), 3),
                    "Recommended (Lowest CO2)": "✓" if opt.get("recommended") else ""
                })
            pd.DataFrame(opt_rows).to_excel(writer, sheet_name="Route Options", index=False)

        # Sheet 4: Emission factors reference
        from modules.emission_factors import get_all_factors_df
        get_all_factors_df().to_excel(writer, sheet_name="Emission Factors", index=False)

        # Style the sheets
        wb = writer.book
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
            # Bold header row
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor="D9E1F2"
                )

    print(f"✓ Excel saved: {fname}")
    return fname


# ---------------------------------------------------------------------------
# MULTI-JOURNEY EXPORT
# ---------------------------------------------------------------------------

def export_multi_journey_csv(multi_result: Dict, output_dir: str = "outputs") -> str:
    """Export multi-journey summary to CSV."""
    from modules.carbon_calculator import multi_results_to_dataframe

    df = multi_results_to_dataframe(multi_result)
    os.makedirs(output_dir, exist_ok=True)
    fname = f"{output_dir}/conference_footprint_{_timestamp_str()}.csv"
    df.to_csv(fname, index=False)
    print(f"✓ CSV saved: {fname}")
    return fname


def export_multi_journey_xlsx(multi_result: Dict, output_dir: str = "outputs") -> str:
    """Export multi-journey results to a comprehensive Excel workbook."""
    from modules.carbon_calculator import multi_results_to_dataframe, results_to_dataframe

    if not OPENPYXL_AVAILABLE:
        print("openpyxl not installed. Run: pip install openpyxl")
        return None

    summary_df = multi_results_to_dataframe(multi_result)
    os.makedirs(output_dir, exist_ok=True)
    fname = f"{output_dir}/conference_footprint_{_timestamp_str()}.xlsx"

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        # Sheet 1: Summary across all journeys
        summary_df.to_excel(writer, sheet_name="All Journeys", index=False)

        # Sheet 2: Totals
        totals = pd.DataFrame([{
            "Total Journeys": multi_result["num_journeys"],
            "Total CO2 (kg CO2e)": multi_result["total_co2_kg"],
            "Average CO2 per Journey (kg)": round(
                multi_result["total_co2_kg"] / max(multi_result["num_journeys"], 1), 1
            ),
            "CO2 in Tonnes": round(multi_result["total_co2_kg"] / 1000, 3),
        }])
        totals.to_excel(writer, sheet_name="Totals", index=False)

        # Sheet 3+: Individual journey details
        for r in multi_result["journeys"][:10]:  # Cap at 10 sheets
            sheet_name = f"J{r['journey_id']}_{r.get('attendee_label','')[:12]}"
            sheet_name = sheet_name[:31].replace("/", "-")
            df = results_to_dataframe(r)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet: Emission factors reference
        from modules.emission_factors import get_all_factors_df
        get_all_factors_df().to_excel(writer, sheet_name="Emission Factors", index=False)

        # Style
        wb = writer.book
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor="D9E1F2"
                )

    print(f"✓ Excel workbook saved: {fname}")
    return fname


def export_json(result_or_multi: Dict, output_dir: str = "outputs") -> str:
    """Export full result dict to JSON (for archiving)."""
    os.makedirs(output_dir, exist_ok=True)
    fname = f"{output_dir}/footprint_{_timestamp_str()}.json"

    # Make serialisable (remove non-serialisable objects)
    def clean(obj):
        if isinstance(obj, dict):
            return {k: clean(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [clean(i) for i in obj]
        if isinstance(obj, (int, float, str, bool, type(None))):
            return obj
        return str(obj)

    with open(fname, "w") as f:
        json.dump(clean(result_or_multi), f, indent=2)
    print(f"✓ JSON saved: {fname}")
    return fname


# ---------------------------------------------------------------------------
# JUPYTER DOWNLOAD HELPER
# ---------------------------------------------------------------------------

def create_download_link(filepath: str, display_text: str = None) -> str:
    """
    Create an HTML download link for use in Jupyter notebooks.
    Usage: from IPython.display import display, HTML
           display(HTML(create_download_link('outputs/file.csv')))
    """
    import base64

    if not os.path.exists(filepath):
        return f"<p>File not found: {filepath}</p>"

    with open(filepath, "rb") as f:
        data = base64.b64encode(f.read()).decode()

    filename = os.path.basename(filepath)
    ext = filename.split(".")[-1].lower()

    mime_types = {
        "csv": "text/csv",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json",
        "html": "text/html",
    }
    mime = mime_types.get(ext, "application/octet-stream")
    text = display_text or f"⬇ Download {filename}"

    return (
        f'<a href="data:{mime};base64,{data}" download="{filename}" '
        f'style="background:#3498db; color:white; padding:8px 16px; '
        f'border-radius:6px; text-decoration:none; font-family:Arial; '
        f'font-size:14px;">{text}</a>'
    )
